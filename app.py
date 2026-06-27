import os
import json
import io
import base64
import requests
import anthropic
import fitz  # pymupdf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=False)


@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    return response


@app.route("/extract", methods=["OPTIONS"])
@app.route("/health", methods=["OPTIONS"])
def handle_options():
    return "", 204


client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))


def render_pdf_page(pdf_url: str, page_number: int) -> str:
    """Download PDF and render the given page (1-indexed) to base64 PNG."""
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; PDFTableExtractor/1.0)",
    }
    r = requests.get(pdf_url, headers=headers, timeout=30, stream=True)
    r.raise_for_status()

    pdf_bytes = r.content
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    page_idx = page_number - 1  # 0-indexed
    if page_idx < 0 or page_idx >= len(doc):
        page_idx = 0

    page = doc[page_idx]

    # Render at 2x zoom for better quality / readability by Claude
    mat = fitz.Matrix(2.0, 2.0)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    png_bytes = pix.tobytes("png")
    doc.close()

    return base64.b64encode(png_bytes).decode("utf-8")


def parse_tables_from_response(text: str):
    try:
        start = text.find("[")
        end = text.rfind("]") + 1
        if start != -1 and end > start:
            return json.loads(text[start:end])
    except Exception:
        pass
    return []


def build_excel(tables):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", start_color="2E7D32")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", start_color="F1F8E9")

    for i, table in enumerate(tables):
        title = (table.get("title") or f"Table {i + 1}")[:31]
        ws = wb.create_sheet(title=title)

        headers = table.get("headers", [])
        rows = table.get("rows", [])

        if not headers and not rows:
            continue

        row_offset = 1
        if title:
            ncols = max(len(headers), 1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            cell = ws.cell(row=1, column=1, value=title)
            cell.font = Font(bold=True, name="Arial", size=12)
            cell.alignment = center
            row_offset = 2

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_offset, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for row_idx, row in enumerate(rows, 1):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_offset + row_idx, column=col_idx, value=value)
                cell.alignment = left
                cell.border = border
                if fill:
                    cell.fill = fill

        # Auto-size columns (skip MergedCell objects)
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                if hasattr(cell, "column_letter"):
                    col_letter = cell.column_letter
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

        ws.freeze_panes = ws.cell(row=row_offset + 1, column=1)

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    return wb


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/extract", methods=["POST"])
def extract():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    pdf_url = data.get("pdf_url")
    page_number = int(data.get("page", 1))

    if not pdf_url:
        return jsonify({"error": "pdf_url is required"}), 400

    # Render the PDF page server-side
    try:
        image_b64 = render_pdf_page(pdf_url, page_number)
    except requests.exceptions.RequestException as e:
        return jsonify({"error": f"Failed to download PDF: {str(e)}"}), 502
    except Exception as e:
        return jsonify({"error": f"Failed to render PDF page: {str(e)}"}), 500

    prompt = """You are an expert at extracting tables from PDF document images.

Analyze this PDF page image and extract ALL tables you can find.

Return ONLY a JSON array (no markdown, no extra text) in this exact format:
[
  {
    "title": "Table title or description (empty string if none)",
    "headers": ["Column 1", "Column 2", "Column 3"],
    "rows": [
      ["row1col1", "row1col2", "row1col3"],
      ["row2col1", "row2col2", "row2col3"]
    ]
  }
]

Rules:
- Extract every table on the page, even partial ones
- Preserve all text exactly as shown
- If a cell spans multiple columns, repeat the value or use empty strings
- If no tables found, return: []
- Numbers: preserve formatting (e.g. "$1,234.56", "42%")
- Empty cells: use empty string ""
- Do NOT include any explanation, only the JSON array"""

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/png",
                            "data": image_b64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )

    response_text = response.content[0].text.strip()
    tables = parse_tables_from_response(response_text)

    if not tables:
        return jsonify({"error": "No tables found on this page"}), 422

    wb = build_excel(tables)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"page_{page_number}_tables.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
